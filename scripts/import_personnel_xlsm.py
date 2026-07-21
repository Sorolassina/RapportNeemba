"""
Import des données sources du classeur "01_Suivi_du_personnel_v13.8.xlsm"
vers des fichiers JSON exploitables par l'application (app/data/personnel/).

On importe UNIQUEMENT les onglets "source" (données saisies), pas les onglets
de calcul/pivot (TCD1, TCD2, Analyse, Analyse 2026, PC2023..PC2026, Grille TCDP,
etc.) : ces derniers sont recalculés par app/analysis.py à partir des données
sources, pour rester à jour automatiquement.

Les noms de colonnes Excel sont conservés tels quels (mêmes libellés) dans les
clés JSON, pour respecter fidèlement les champs des fichiers d'origine. Les
retours à la ligne dans les en-têtes sont remplacés par un espace simple pour
faciliter l'usage en Python/Jinja, mais le libellé reste identique sinon.

Usage:
    python import_personnel_xlsm.py <chemin_vers_le_fichier.xlsm> <dossier_sortie>
"""
import json
import sys
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")


def _clean_header(h):
    if h is None:
        return None
    return str(h).replace("\n", " ").strip()


def _clean_value(v):
    import datetime as _dt
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.isoformat()
    return v


def _find_header_row(ws, max_scan=10):
    """Trouve la ligne d'en-tête : celle qui a le plus de cellules texte."""
    best_row, best_count = 1, -1
    for r in range(1, min(max_scan, ws.max_row) + 1):
        vals = [c.value for c in ws[r]]
        cnt = sum(1 for v in vals if isinstance(v, str) and v.strip())
        if cnt > best_count:
            best_count, best_row = cnt, r
    return best_row


def sheet_to_records(ws, header_row=None):
    """Convertit une feuille en liste de dicts {en-tête: valeur}."""
    if header_row is None:
        header_row = _find_header_row(ws)

    headers = [_clean_header(c.value) for c in ws[header_row]]
    records = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None for v in row):
            continue
        rec = {}
        has_data = False
        for h, v in zip(headers, row):
            if h is None:
                continue
            cv = _clean_value(v)
            rec[h] = cv
            if cv not in (None, ""):
                has_data = True
        if has_data:
            records.append(rec)
    return {"header_row": header_row, "headers": [h for h in headers if h], "records": records}


# Onglets "sources" à importer (on laisse de côté les onglets de calcul/pivot :
# TCD1, TCD2, Analyse, Analyse 2026, PC2023, PC2024, PC2025, PC2026,
# Grille TCDP, 1. F. English, 2. DPC Ass., 2. ILT Neemba, 3. Certif.,
# 5. ILT CAT, Autres marques, Autres formations, Form Spé, Feuil1, TdB1, TdB2,
# Base3, Poste TCDP, Code Société, Setting, Sommaire —
# ces derniers seront recalculés dans app/analysis.py).
SHEETS_TO_IMPORT = {
    "Base 1": "base1_personnel.json",
    "Base 2": "base2_avancement.json",
    "Recrutement": "recrutement.json",
    "Départ1": "departs.json",
    "Objectif": "objectifs.json",
    "Cap&Cap": "cap_and_cap.json",
}


def main():
    if len(sys.argv) < 3:
        print("Usage: python import_personnel_xlsm.py <fichier.xlsm> <dossier_sortie>")
        sys.exit(1)

    src = Path(sys.argv[1])
    out_dir = Path(sys.argv[2])
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(src, data_only=True, keep_vba=False)

    summary = {}
    for sheet_name, out_file in SHEETS_TO_IMPORT.items():
        if sheet_name not in wb.sheetnames:
            print(f"[!] Onglet introuvable, ignoré : {sheet_name}")
            continue
        ws = wb[sheet_name]
        data = sheet_to_records(ws)
        out_path = out_dir / out_file
        out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        summary[sheet_name] = {
            "fichier": out_file,
            "header_row": data["header_row"],
            "nb_champs": len(data["headers"]),
            "nb_lignes": len(data["records"]),
        }
        print(f"[ok] {sheet_name} -> {out_file} "
              f"({len(data['records'])} lignes, {len(data['headers'])} champs, "
              f"en-tête ligne {data['header_row']})")

    (out_dir / "_import_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print("\nRésumé :")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
