"""
Extrait TOUS les onglets des deux classeurs Excel ("01_Suivi_du_personnel" et
"02_Suivi du tutorat") sous forme de grilles brutes (valeurs cellule par
cellule, comme dans Excel), pour garantir qu'aucune vue ne manque dans
l'application — chaque onglet du classeur devient une page consultable dans
le volet Techniciens ("Vues Excel").

Contrairement aux imports "structurés" (Base 1, Base 2, Recrutement,
Objectif, Cap&Cap, OJT, Référentiel — qui alimentent les fonctionnalités
CRUD du module Techniciens), ce script ne fait aucune hypothèse sur les
en-têtes : il recopie la grille telle quelle (lignes/colonnes Excel), pour
un usage de consultation / comparaison fidèle avec le fichier source.

Usage :
    python scripts/dump_excel_views.py "<fichier.xlsm ou .xlsx>" <dossier_sortie> <fichier_label>
"""
import json
import sys
import warnings
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

MAX_ROWS = 3000
MAX_COLS = 1100


def _clean_value(v):
    import datetime as _dt
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.isoformat()
    return v


def sheet_to_grid(ws):
    scan_rows = min(ws.max_row or 1, MAX_ROWS)
    scan_cols = min(ws.max_column or 1, MAX_COLS)
    truncated = (ws.max_row or 0) > MAX_ROWS or (ws.max_column or 0) > MAX_COLS

    raw = []
    for row in ws.iter_rows(min_row=1, max_row=scan_rows, max_col=scan_cols, values_only=True):
        raw.append([_clean_value(v) for v in row])

    # trim colonnes vides en fin de grille
    last_col = 0
    for r in raw:
        for i in range(len(r) - 1, -1, -1):
            if r[i] not in (None, ""):
                last_col = max(last_col, i + 1)
                break
    last_col = last_col or 1

    # trim lignes vides en fin de grille
    last_row = 0
    for i in range(len(raw) - 1, -1, -1):
        if any(v not in (None, "") for v in raw[i][:last_col]):
            last_row = i + 1
            break
    last_row = last_row or 1

    rows = [r[:last_col] for r in raw[:last_row]]
    col_letters = [get_column_letter(c) for c in range(1, last_col + 1)]
    return {
        "nb_lignes": last_row,
        "nb_colonnes": last_col,
        "colonnes": col_letters,
        "rows": rows,
        "truncated": truncated,
    }


def main():
    if len(sys.argv) < 4:
        print("Usage: python dump_excel_views.py <fichier> <dossier_sortie> <fichier_label>")
        sys.exit(1)

    src = Path(sys.argv[1])
    out_dir = Path(sys.argv[2])
    label = sys.argv[3]
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(src, data_only=True, keep_vba=False)

    index = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        grid = sheet_to_grid(ws)
        safe_name = sheet_name.replace("/", "-")
        out_path = out_dir / f"{safe_name}.json"
        out_path.write_text(json.dumps(grid, ensure_ascii=False), encoding="utf-8")
        index.append({
            "nom": sheet_name, "fichier": f"{safe_name}.json",
            "nb_lignes": grid["nb_lignes"], "nb_colonnes": grid["nb_colonnes"],
            "truncated": grid["truncated"],
        })
        print(f"[ok] {sheet_name} -> {safe_name}.json "
              f"({grid['nb_lignes']}x{grid['nb_colonnes']}{' TRONQUÉ' if grid['truncated'] else ''})")

    (out_dir / "_index.json").write_text(
        json.dumps({"label": label, "sheets": index}, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"\n{len(index)} onglets exportés pour '{label}'.")


if __name__ == "__main__":
    main()
