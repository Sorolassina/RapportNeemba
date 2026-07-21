"""
Importe TOUS les onglets restants (ceux qui n'ont pas encore de store CRUD
dédié : Base 1, Base 2, Recrutement, Objectif, Cap&Cap, Départ1, 7. OJT (1),
Référentiel de compétences) sous forme de tables éditables génériques, afin
que chaque onglet des deux classeurs soit manipulable en CRUD dans
l'application — pas seulement consultable.

Contrairement à scripts/dump_excel_views.py (grille brute en lecture seule,
abandonnée), ce script détecte une ligne d'en-tête par onglet et produit une
vraie liste d'enregistrements {champ: valeur, id: ...}, exploitable par
app/onglets_store.py pour du create/update/delete générique.

Usage :
    python scripts/import_remaining_sheets.py
"""
import json
import uuid
import warnings
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "app" / "data" / "onglets"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Onglets déjà couverts par un store CRUD dédié -> pas ré-importés ici
DEJA_COUVERTS_PERSONNEL = {
    "Base 1", "Base 2", "Recrutement", "Objectif", "Cap&Cap", "Départ1",
}
DEJA_COUVERTS_TUTORAT = {"7. OJT (1)", "Référentiel de compétences"}

MAX_ROWS = 2000
MAX_COLS = 200


def _clean_header(h):
    if h is None:
        return None
    return str(h).replace("\n", " ").strip()


def _clean_value(v):
    import datetime as _dt
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.isoformat()
    return v


def _find_header_row(ws, max_scan=12):
    best_row, best_count = 1, -1
    for r in range(1, min(max_scan, ws.max_row or 1) + 1):
        vals = [c.value for c in ws[r][:MAX_COLS]]
        cnt = sum(1 for v in vals if isinstance(v, str) and v.strip())
        if cnt > best_count:
            best_count, best_row = cnt, r
    return best_row, best_count


def sheet_to_records(ws):
    header_row, score = _find_header_row(ws)
    raw_headers = [c.value for c in ws[header_row][:MAX_COLS]]

    if score >= 2:
        headers = []
        seen = {}
        for h in raw_headers:
            h = _clean_header(h)
            if not h:
                headers.append(None)
                continue
            if h in seen:
                seen[h] += 1
                h = f"{h} ({seen[h]})"
            else:
                seen[h] = 0
            headers.append(h)
        data_start = header_row + 1
    else:
        # Pas d'en-tête détectable (onglet pivot/résumé) : colonnes génériques
        # A, B, C... et on garde toutes les lignes comme données.
        max_col = min(ws.max_column or 1, MAX_COLS)
        headers = [get_column_letter(c) for c in range(1, max_col + 1)]
        data_start = 1

    records = []
    max_row = min(ws.max_row or 1, MAX_ROWS)
    for row in ws.iter_rows(min_row=data_start, max_row=max_row, max_col=len(headers), values_only=True):
        if all(v is None for v in row):
            continue
        rec = {}
        has_data = False
        for h, v in zip(headers, row):
            if h is None:
                continue
            cv = _clean_value(v)
            rec[h] = cv
            if cv not in (None, ""):
                has_data = True
        if has_data:
            rec["id"] = uuid.uuid4().hex[:10]
            records.append(rec)

    return {
        "headers": [h for h in headers if h],
        "header_detecte": score >= 2,
        "header_row": header_row,
        "records": records,
    }


def process_workbook(path, already_covered, out_subdir, label):
    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=False)
    out_dir = OUT_DIR / out_subdir
    out_dir.mkdir(parents=True, exist_ok=True)
    index = []
    for sheet_name in wb.sheetnames:
        if sheet_name in already_covered:
            continue
        ws = wb[sheet_name]
        data = sheet_to_records(ws)
        safe = sheet_name.replace("/", "-")
        out_path = out_dir / f"{safe}.json"
        out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        index.append({
            "nom": sheet_name, "fichier": f"{safe}.json",
            "nb_champs": len(data["headers"]), "nb_lignes": len(data["records"]),
            "header_detecte": data["header_detecte"],
        })
        print(f"[ok] {label} / {sheet_name} -> {safe}.json "
              f"({len(data['records'])} lignes, {len(data['headers'])} champs"
              f"{'' if data['header_detecte'] else ', SANS EN-TETE (colonnes génériques)'})")
    (out_dir / "_index.json").write_text(
        json.dumps({"label": label, "sheets": index}, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return index


def main():
    idx1 = process_workbook(
        ROOT / "01_Suivi_du_personnel_v13.8 (1).xlsm",
        DEJA_COUVERTS_PERSONNEL, "personnel", "Suivi du personnel",
    )
    idx2 = process_workbook(
        ROOT / "02_Suivi du tutorat_v2 (1).xlsx",
        DEJA_COUVERTS_TUTORAT, "tutorat", "Suivi du tutorat",
    )
    print(f"\n{len(idx1)} onglets (personnel) + {len(idx2)} onglet(s) (tutorat) importés en CRUD générique.")


if __name__ == "__main__":
    main()
