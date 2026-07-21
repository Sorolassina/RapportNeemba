"""
Import des données du classeur "02_Suivi du tutorat_v2.xlsx" (OJT + référentiel
de compétences) vers des fichiers JSON exploitables par l'application.

Onglets traités :
- "7. OJT (1)" : matrice de progression OJT, une ligne par technicien
  (identifié par User Id / Mtle), avec un champ par compétence
  (codes du type NEE/07/041) + les infos de fiche (poste, niveau atelier...).
- "Référentiel de compétences" : grille de référence (code compétence,
  description, niveau requis par famille de machine).

Les libellés de colonnes Excel sont conservés tels quels.

Usage:
    python import_tutorat_xlsx.py <chemin_vers_le_fichier.xlsx> <dossier_sortie>
"""
import json
import sys
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")


def _clean_header(h):
    if h is None:
        return None
    return str(h).replace("\n", " ").strip()


def _clean_value(v):
    import datetime as _dt
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.isoformat()
    return v


# Colonnes "fiche" fixes en tête de la matrice OJT (avant les colonnes de
# compétences). Après ces colonnes, on trouve les niveaux ateliers (N2A..N4)
# puis une colonne par code de compétence (NEE/xx/xxx).
OJT_META_COLS = [
    "User Id", "Code société", "Mtle", "Nom / Prénom", "Statut", "Service",
    "Programme", "Poste harmonisé actuel", "Poste harmonisé cible",
    "SGA Fait", "Date d'entrée", "Niveau atelier actuel", "Niveau atelier cible",
]


def import_ojt(wb, out_dir):
    sheet_name = "7. OJT (1)"
    if sheet_name not in wb.sheetnames:
        print(f"[!] Onglet introuvable : {sheet_name}")
        return
    ws = wb[sheet_name]

    # La ligne d'en-tête réelle est celle qui contient "User Id" en colonne A
    header_row = None
    for r in range(1, min(20, ws.max_row) + 1):
        if ws.cell(row=r, column=1).value == "User Id":
            header_row = r
            break
    if header_row is None:
        print(f"[!] Ligne d'en-tête introuvable dans {sheet_name}")
        return

    headers = [_clean_header(ws.cell(row=header_row, column=c).value)
               for c in range(1, ws.max_column + 1)]

    # Découpage : colonnes meta / colonnes niveaux (N2A..N4, sans "/") / colonnes compétences (contiennent "/")
    meta_idx, niveau_idx, comp_idx = [], [], []
    for i, h in enumerate(headers):
        if h is None:
            continue
        if h in OJT_META_COLS:
            meta_idx.append(i)
        elif "/" in h:
            comp_idx.append(i)
        else:
            niveau_idx.append(i)

    records = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[0] is None:  # pas de User Id -> ligne vide
            continue
        rec = {"fiche": {}, "niveaux": {}, "competences": {}}
        for i in meta_idx:
            rec["fiche"][headers[i]] = _clean_value(row[i]) if i < len(row) else None
        for i in niveau_idx:
            v = row[i] if i < len(row) else None
            if v not in (None, "", 0):
                rec["niveaux"][headers[i]] = _clean_value(v)
        for i in comp_idx:
            v = row[i] if i < len(row) else None
            if v not in (None, "", 0):
                rec["competences"][headers[i]] = _clean_value(v)
        records.append(rec)

    out = {
        "header_row": header_row,
        "nb_competences_suivies": len(comp_idx),
        "records": records,
    }
    out_path = out_dir / "ojt_progression.json"
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[ok] {sheet_name} -> ojt_progression.json "
          f"({len(records)} techniciens, {len(comp_idx)} compétences suivies)")


def import_referentiel(wb, out_dir):
    sheet_name = "Référentiel de compétences"
    if sheet_name not in wb.sheetnames:
        print(f"[!] Onglet introuvable : {sheet_name}")
        return
    ws = wb[sheet_name]

    header_row = None
    for r in range(1, min(10, ws.max_row) + 1):
        if ws.cell(row=r, column=1).value == "Code compétence":
            header_row = r
            break
    if header_row is None:
        print(f"[!] Ligne d'en-tête introuvable dans {sheet_name}")
        return

    headers = [_clean_header(c.value) for c in ws[header_row]]
    records = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[0] is None:
            continue
        rec = {h: _clean_value(v) for h, v in zip(headers, row) if h}
        records.append(rec)

    out = {"header_row": header_row, "headers": [h for h in headers if h], "records": records}
    out_path = out_dir / "referentiel_competences.json"
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[ok] {sheet_name} -> referentiel_competences.json ({len(records)} compétences)")


def main():
    if len(sys.argv) < 3:
        print("Usage: python import_tutorat_xlsx.py <fichier.xlsx> <dossier_sortie>")
        sys.exit(1)

    src = Path(sys.argv[1])
    out_dir = Path(sys.argv[2])
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(src, data_only=True)
    import_ojt(wb, out_dir)
    import_referentiel(wb, out_dir)


if __name__ == "__main__":
    main()
